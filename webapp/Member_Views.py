import csv
import io
import logging

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages

from app.models import MembershipPayment, Alumni
from webapp.sheets_sync import sync_all, _normalize_header, parse_row_to_fields

logger = logging.getLogger(__name__)


@login_required(login_url='/login')
def payment_records_view(request):
    """HOD only: full payment records table with sync/upload controls."""
    if request.user.user_type != '1':
        messages.error(request, 'Access denied.')
        return redirect('landing')

    batch = request.GET.get('batch', '')
    status = request.GET.get('status', '')
    search = request.GET.get('search', '')

    qs = MembershipPayment.objects.all()
    if batch:
        qs = qs.filter(batch_year=batch)
    if status:
        qs = qs.filter(status=status)
    if search:
        qs = qs.filter(member_name__icontains=search)

    life_members = qs.filter(member_type='LIFE')
    general_members = qs.filter(member_type='GENERAL')
    life_count = MembershipPayment.objects.filter(member_type='LIFE').count()
    general_count = MembershipPayment.objects.filter(member_type='GENERAL').count()
    batch_years = (
        MembershipPayment.objects.values_list('batch_year', flat=True)
        .distinct().exclude(batch_year='').order_by('batch_year')
    )
    last_synced = (
        MembershipPayment.objects.order_by('-last_synced_at')
        .values_list('last_synced_at', flat=True).first()
    )

    return render(request, 'members/payment_records.html', {
        'life_members': life_members,
        'general_members': general_members,
        'total_count': life_count + general_count,
        'life_count': life_count,
        'general_count': general_count,
        'batch_years': batch_years,
        'last_synced': last_synced,
        'filter_batch': batch,
        'filter_status': status,
        'search': search,
    })


@login_required(login_url='/login')
def trigger_sync_view(request):
    """HOD only: manually trigger Google Sheets sync."""
    if request.user.user_type != '1':
        messages.error(request, 'Access denied.')
        return redirect('landing')

    result = sync_all()
    if result['success']:
        messages.success(
            request,
            f"Sync complete! Added: {result['total_added']}, "
            f"Updated: {result['total_updated']}, Errors: {result['total_errors']}"
        )
    else:
        messages.error(request, f"Sync failed: {result.get('error', 'Unknown error')}")

    return redirect('payment_records')


@login_required(login_url='/login')
def export_payments_csv(request):
    """HOD only: export payment records as CSV."""
    if request.user.user_type != '1':
        messages.error(request, 'Access denied.')
        return redirect('landing')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="chandramonians_payments.csv"'

    writer = csv.writer(response)
    writer.writerow(['Name', 'Type', 'Batch Year', 'Payment Date', 'Amount', 'Receipt No', 'Phone', 'Status'])

    for p in MembershipPayment.objects.all().order_by('member_type', 'member_name'):
        writer.writerow([
            p.member_name, p.get_member_type_display(), p.batch_year,
            p.payment_date or '', p.amount or '', p.receipt_number,
            p.phone, p.get_status_display(),
        ])

    return response


# ─── File Upload Helpers ────────────────────────────────────────────────────

def _upsert_rows(rows_iter, member_type):
    """Upsert an iterable of {normalised_header: value} dicts. Returns stats dict."""
    stats = {'added': 0, 'updated': 0, 'skipped': 0, 'errors': 0}
    for row_idx, row_data in rows_iter:
        try:
            if not any(v for v in row_data.values() if str(v).strip()):
                continue
            payment_fields = parse_row_to_fields(row_data)
            if not payment_fields.get('member_name'):
                stats['skipped'] += 1
                continue
            _, created = MembershipPayment.objects.update_or_create(
                member_type=member_type,
                sheet_row_id=row_idx,
                defaults=payment_fields,
            )
            if created:
                stats['added'] += 1
            else:
                stats['updated'] += 1
        except Exception as e:
            logger.error(f"Row {row_idx} ({member_type}): {e}")
            stats['errors'] += 1
    return stats


def _iter_csv(decoded_text):
    """Yield (row_idx, normalised_row_dict) from CSV text."""
    reader = csv.DictReader(io.StringIO(decoded_text))
    for row_idx, row in enumerate(reader, start=2):
        yield row_idx, {_normalize_header(k): v for k, v in row.items()}


def _iter_worksheet(ws):
    """Yield (row_idx, normalised_row_dict) from an openpyxl worksheet."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return
    headers = [_normalize_header(str(c) if c is not None else '') for c in rows[0]]
    for row_idx, row in enumerate(rows[1:], start=2):
        yield row_idx, dict(zip(headers, [str(c) if c is not None else '' for c in row]))


# ─── CSV Upload ──────────────────────────────────────────────────────────────

@login_required(login_url='/login')
def upload_payments_csv_view(request):
    """HOD only: import one member-type tab from an uploaded CSV file."""
    if request.user.user_type != '1':
        messages.error(request, 'Access denied.')
        return redirect('landing')

    if request.method != 'POST':
        return redirect('payment_records')

    csv_file = request.FILES.get('csv_file')
    member_type = request.POST.get('member_type', 'LIFE')

    if member_type not in ('LIFE', 'GENERAL'):
        messages.error(request, 'Invalid member type.')
        return redirect('payment_records')
    if not csv_file:
        messages.error(request, 'No file selected.')
        return redirect('payment_records')
    if not csv_file.name.lower().endswith('.csv'):
        messages.error(request, 'Please upload a .csv file.')
        return redirect('payment_records')

    try:
        decoded = csv_file.read().decode('utf-8-sig')  # utf-8-sig strips Excel BOM
        stats = _upsert_rows(_iter_csv(decoded), member_type)
        messages.success(
            request,
            f"CSV imported ({member_type.title()} Members)! "
            f"Added: {stats['added']}, Updated: {stats['updated']}, "
            f"Skipped: {stats['skipped']}, Errors: {stats['errors']}"
        )
    except Exception as e:
        messages.error(request, f"CSV import failed: {e}")

    return redirect('payment_records')


# ─── Excel Upload ─────────────────────────────────────────────────────────────

@login_required(login_url='/login')
def upload_payments_excel_view(request):
    """HOD only: import both member tabs from a single Excel (.xlsx) file.

    Downloads from Google Sheets via File → Download → Microsoft Excel (.xlsx)
    will contain all tabs. Tabs whose name contains 'life' map to LIFE members;
    tabs containing 'general' map to GENERAL members.
    """
    if request.user.user_type != '1':
        messages.error(request, 'Access denied.')
        return redirect('landing')

    if request.method != 'POST':
        return redirect('payment_records')

    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        messages.error(request, 'No file selected.')
        return redirect('payment_records')
    if not excel_file.name.lower().endswith('.xlsx'):
        messages.error(request, 'Please upload a .xlsx file (Excel format).')
        return redirect('payment_records')

    try:
        import openpyxl
        wb = openpyxl.load_workbook(
            io.BytesIO(excel_file.read()), read_only=True, data_only=True
        )
    except Exception as e:
        messages.error(request, f'Could not open Excel file: {e}')
        return redirect('payment_records')

    totals = {'added': 0, 'updated': 0, 'skipped': 0, 'errors': 0}
    sheets_processed = []

    for sheet_name in wb.sheetnames:
        name_lower = sheet_name.lower()
        if 'life' in name_lower:
            member_type = 'LIFE'
        elif 'general' in name_lower:
            member_type = 'GENERAL'
        else:
            logger.info(f"Skipping unrecognised tab: {sheet_name}")
            continue

        stats = _upsert_rows(_iter_worksheet(wb[sheet_name]), member_type)
        for k in totals:
            totals[k] += stats[k]
        sheets_processed.append(f'"{sheet_name}" → {member_type}')

    wb.close()

    if not sheets_processed:
        messages.error(
            request,
            'No recognised tabs found. Tab names must contain "Life" or "General".'
        )
        return redirect('payment_records')

    messages.success(
        request,
        f"Excel imported! Tabs: {', '.join(sheets_processed)}. "
        f"Added: {totals['added']}, Updated: {totals['updated']}, "
        f"Skipped: {totals['skipped']}, Errors: {totals['errors']}"
    )
    return redirect('payment_records')


# ─── Alumni & Member Views ────────────────────────────────────────────────────

@login_required(login_url='/login')
def alumni_own_payment_view(request):
    """Alumni: view only their own payment record."""
    if request.user.user_type != '3':
        messages.error(request, 'Access denied.')
        return redirect('landing')

    try:
        alumni = request.user.alumni
        full_name = f"{request.user.first_name} {request.user.last_name}".strip()
        payment = MembershipPayment.objects.filter(
            member_name__icontains=full_name.split()[0]
        ).first() if full_name else None
    except Exception:
        payment = None
        alumni = None

    return render(request, 'members/my_payment.html', {
        'payment': payment,
        'alumni': alumni,
    })


@login_required(login_url='/login')
def alumni_directory_view(request):
    """Full alumni directory for logged-in alumni."""
    batch = request.GET.get('batch', '')
    profession = request.GET.get('profession', '')
    search = request.GET.get('search', '')

    qs = Alumni.objects.select_related('admin').all()
    if batch:
        qs = qs.filter(passing_year=batch)
    if profession:
        qs = qs.filter(profession__icontains=profession)
    if search:
        qs = qs.filter(admin__first_name__icontains=search) | qs.filter(admin__last_name__icontains=search)

    batch_years = (
        Alumni.objects.values_list('passing_year', flat=True)
        .distinct().exclude(passing_year=None).order_by('passing_year')
    )

    return render(request, 'members/directory.html', {
        'alumni': qs,
        'batch_years': batch_years,
        'filter_batch': batch,
        'search': search,
    })


@login_required(login_url='/login')
def job_board_view(request):
    """Placeholder: job board for alumni networking."""
    return render(request, 'members/job_board.html', {})


@login_required(login_url='/login')
def mentorship_view(request):
    """Alumni can sign up as mentors."""
    if request.method == 'POST':
        messages.success(request, 'Thank you for signing up as a mentor! We will contact you soon.')
        return redirect('mentorship')
    return render(request, 'members/mentorship.html', {})
